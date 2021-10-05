import selenium
import pytest
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.webdriver.chrome import options
from selenium.webdriver import DesiredCapabilities
from selenium.webdriver.common import actions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import selenium.common.exceptions as Sele_exception 
import csv
import os
import time
import logging
import openpyxl
from datetime import datetime
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
import pytest_html
import pytest_html_reporter

# @pytest.mark.usefixtures("setup")
# exe_dict = {}
class Test_Selenium_Execution():
    global exe_list, exe_dict2, credential, logger, fhandler, line_exe
    line_exe = 0
    credential = []
    try:
        f_val = open("login_credits.txt","r")
        credential = (f_val.read()).split(",")
        f_val.close()
    except FileNotFoundError:
        for credits in range(0,7):
            credential.append("")
    exe_dict2 = {}
    fileread = open("Run_cases.csv",'r')
    exe_list = (fileread.read()).split("\n")
    fileread.close()
    exe_list.pop(0)
    exe_list.pop(len(exe_list)-1)
    for i in range(0,len(exe_list)):
        exe_dict2[exe_list[i]] = exe_list[i]
    # print(exe_dict2)
    logger = logging.getLogger()
    fhandler = logging.FileHandler("logfile.log")        
    logger.setLevel(logging.INFO)
    # logger = logging.getLogger()
# fhandler = logging.FileHandler("logfile.log")
# logger.setLevel(logging.INFO)
    filename = ""
    fail_name = ""
    # exe_dict2 = ""  
    def wait_till(self,wait_time):
        time.sleep(wait_time)
    def singleaction(self,element,action,values):
        # self.wait_till(1)
        if "click" in action:
            element.click()
        elif "input" in action:
            element.clear()
            element.send_keys(values)
    def multiaction(self,element,action,values):
        if "singleclick" in action and len(values) > 0:
            element[int(values)-1].click()
        else:
            for ele in element:
                if "click" in action:
                    ele.click()
                elif "input" in action:
                    ele.clear()
                    ele.send_keys(values)
    def byeid(self,element,action,values):
        ele = self.driver.find_element(By.ID, element)
        self.singleaction(ele,action,values)        
    def bycss(self,element,action,values):        
        ele = self.driver.find_element(By.CSS_SELECTOR, element)
        if action == "assert":
            self.assertion(ele,action,values)
        else:
            self.singleaction(ele,action,values)        
    def byxpath(self,element,action,values):
        ele = self.driver.find_element(By.XPATH, element)
        if action == "assert":
            self.assertion(ele,action,values)
        else:
            self.singleaction(ele,action,values)      
    def bymultixpath(self,element,action,values):
        ele = self.driver.find_elements_by_xpath(element)
        self.multiaction(ele,action,values)
    def bymulticss(self,element,action,values):
        ele = self.driver.find_elements_by_css_selector(element)
        self.multiaction(ele,action,values)
    def byplink(self,element,action,values):
        ele = self.driver.find_element(By.PARTIAL_LINK_TEXT, element)
        self.singleaction(ele,action,values)        
    def byelink(self,element,action,values):
        ele = self.driver.find_element(By.LINK_TEXT, element)
        self.singleaction(ele,action,values) 
    def byswitchframe(self,element,action,values):
        if "frame" in element:
            self.driver.switch_to.frame(int(values))
    def capturealert(self,element,action,values):
        alert_mode = self.driver.switch_to.alert
        if action in "ok":
            alert_mode.accept()
        if action in "cancel":
            alert_mode.dismiss()
        # driver.switch_to_default_content()
    def switchwindowto(self,element,action,values):
        window_no = self.driver.window_handles[int(values)-1]
        self.driver.switch_to.window(window_no)
    def py_wait(self,element,action,values):
        if "pywait" in action:
            self.wait_till(int(values))
        elif "expectedwait" in action:
            WebDriverWait(self.driver, int(values)).until(EC.presence_of_element_located((By.XPATH, element)))
        elif "extendwait" in action:
            self.driver.implicitly_wait(int(values))
    def saveimage(self,element,action,values):
        now = datetime.now()
        dt_string = now.strftime("%d.%m.%Y %H-%M-%S")
        dt_string2 = now.strftime("%d.%m.%Y")
        # Use to create a folder named screenshot for custom and failed execution screenshots
        filedirect = os.path.dirname(os.path.realpath('__file__'))
        filedirect = os.path.join(filedirect,"screenshot")
        if not os.path.isdir(filedirect):
            os.mkdir(filedirect)
        filedirect = os.path.join(filedirect,"Date-"+str(dt_string2))
        if not os.path.isdir(filedirect):
            os.mkdir(filedirect)
            filedirect = os.path.join(filedirect,"Fail_snaps")
            os.mkdir(filedirect)
        if "fail_error" not in str(element):
            self.driver.save_screenshot("screenshot/Date-"+str(dt_string2)+"/"+str(element)+" - "+str(dt_string)+".png")
        elif "fail_error" in str(element):
            self.driver.save_screenshot("screenshot/Date-"+str(dt_string2)+"/Fail_snaps/"+str(element)+" - "+str(fail_name.replace("/","-"))+" - "+str(dt_string)+".png")
    def bypid(self,element,action,values):
        ele = self.driver.find_element_by_css_selector("[id *='"+str(element)+"']")
        self.singleaction(ele,action,values)
    def destroywindow(self,element,action,values):
        self.driver.close()
    def jsscript(self,element,action,values):
        self.driver.execute_script(element)
    def browser_refresh(self,element,action,values):
        self.driver.refresh()
    def move_to_click(self,element,action,values):
        ele = self.driver.find_element_by_xpath(element)
        self.driver.execute_script("arguments[0].scrollIntoView();",ele)
        ActionChains(self.driver).move_to_element(ele).click().perform()
    def scrollbar(self,element,action,values):
        ele = self.driver.find_element_by_xpath(element)
        self.driver.execute_script("arguments[0].scrollIntoView();",ele)
    def log_credits(self):
        if len(credential) == 6:
            if credential[6] == "logit":
                logger.info("USERNAME : "+credential[0])
                logger.info("PASSWORD : "+credential[1])
                logger.info("ENV : "+credential[2])
            else:
                pass
        else:
            pass
    def init_web(self):
        opt = Options()
        if "min" in credential[5]:
            opt.headless = True
            logger.info("Executing Scripts in Headless mode")
        self.driver = webdriver.Chrome(options=opt)
        self.driver.set_window_size(1500, 820)
        self.driver.implicitly_wait(5)
    def old_initialise(self):
        self.init_web()
        self.driver.get(credential[2])
        self.driver.find_element(By.ID, "logincheck_username").send_keys(credential[0])
        self.driver.find_element(By.ID, "logincheck_password").send_keys(credential[1])
        self.driver.find_element(By.ID, "logincheck_password").send_keys(Keys.ENTER)
        if 5 < len(credential) and int(credential[3]) != 0:
            try:
                combo = self.driver.find_element_by_css_selector("[name='centername']")
                combodata = self.driver.find_elements_by_xpath('//select[@name="centername"]/option')
                comval = []
                for i in combodata:
                    comval.append(self.driver.execute_script("return arguments[0].value",i))
                    # print(comval)
                if int(credential[3]) < len(comval):
                    combo.find_element_by_css_selector("option[value='"+comval[int(credential[3])]+"']").click()
                    self.driver.find_element(By.NAME, "SignIN").click()
                    self.wait_till(3)
                else:
                    combo.find_element_by_css_selector("option[value='"+comval[1]+"']").click()
                    self.driver.find_element(By.NAME, "SignIN").click()
            except NoSuchElementException:
                pass
        else:
            pass  
        self.log_credits()
    def assertion(self,element,action,values):
        # ele = self.driver.find_element_by_xpath(element)
        check_text = str(values).replace("\\n","")
        verify_text = str(element.text).replace("\n","")
        # print(verify_text,"\n\n",check_text)
        if verify_text in check_text:
            logger.info("Assertion matched :"+values)
            assert verify_text in check_text
        else:
            logger.error("Assertion not matched :"+values)
            assert verify_text in check_text
    def new_initialise(self):
        self.init_web()
        self.driver.get(credential[2])
        self.driver.find_element(By.ID, "mat-input-0").send_keys(credential[0])
        self.driver.find_element(By.ID, "mat-input-1").send_keys(credential[1])
        self.driver.find_element(By.ID, "mat-input-1").send_keys(Keys.ENTER)
        self.wait_till(5)
        if int(credential[3]) == 0:
            pass
        elif int(credential[3]) != 0:
            self.driver.find_element_by_css_selector("span[class *='mat-select-placeholder']").click()
            centercombo = self.driver.find_elements_by_css_selector("span[class *='mat-option-text']")
            centercombo[int(credential[3])-1].click()
            self.wait_till(1)
            # self.driver.execute_script("var a = $('button'); a[2].click();")
            # self.driver.find_element_by_xpath("(//button[contains(text(),'Sign')])[2]").click()
            center_select = self.driver.find_elements_by_css_selector("button[class *='btn']")
            # abcd = self.driver.find_elements_by_xpath("//button")
            self.multiaction(center_select,"singleclick","2")
        self.log_credits()
    def fs_initialise(self):
        self.init_web()
        self.driver.get(credential[2])
        self.driver.find_element(By.CSS_SELECTOR, "[type='text']").send_keys(credential[0])
        self.driver.find_element(By.CSS_SELECTOR, '[type="password"]').send_keys(credential[1])
        self.driver.find_element(By.CSS_SELECTOR, '[type="password"]').send_keys(Keys.ENTER)
        self.wait_till(3)
        self.log_credits()
    def enddriver(self,exe_name,types):
        if types == 0:
            logger.info(str(exe_name)+" : Execution Ended successfully")
        elif types == 1:
            logger.error(str(exe_name)+" : Due to Error, Execution stopped")
        self.driver.quit()
    def logs(self,exe_name):
        global logger
        form = logging.Formatter("%(asctime)s\t%(levelname)s\t"+str("File name\t"+str(exe_name.replace("/","-"))+"\tMessage\t%(message)s"))
        fhandler.setFormatter(form)
        logger.addHandler(fhandler)
        return logger
    def openlink(self,element,action,values):
        if "open" in element:
            self.driver.get(values)
    def getloop(self,element,action,values):
        global line_exe
        line_exe = int(element) - 2
    @pytest.mark.parametrize('exe_dict',exe_dict2)
    def test_execution(self,exe_dict):
        global filename, fail_name, credential
        filename = ""
        fail_name = exe_dict
        exe_log = self.logs(exe_dict)
        xl = 2
        for credits in range(0,7):
            if len(credential) < 4:
                credential.append("")
            elif len(credential) == 0:
                break
            pass
        try:
            # OLD - Old version gingafleets NEW - New version gingafleets FS - Fire safety
            credential[4] = credential[4].strip()
            if credential[4] == "old":
                self.old_initialise()
            elif credential[4] == "new":
                self.new_initialise()
            elif credential[4] == "fs":
                self.fs_initialise()
            else:
                self.init_web()
            exe_lines = []
            fileDir = os.path.dirname(os.path.realpath('__file__'))
            if ".csv" in exe_dict:
                filename = os.path.join(fileDir, "casefolder\\"+exe_dict)
                xl = 0
            elif ".xlsx" in exe_dict:
                filename = os.path.join(fileDir, "casefolder\\"+str(exe_dict))
                xl = 1
            else:
                filename = os.path.join(fileDir, "casefolder\\"+str(exe_dict)+".xlsx")
                xl = 1
            if xl == 1:
                # -------- excel file read -----------
                excel_book = openpyxl.load_workbook(filename)
                excel_sheet = excel_book.active
                for row_i in range(1,excel_sheet.max_row+1):
                    col_val = []
                    for col_j in range(1,excel_sheet.max_column+1):
                        col_val.append(excel_sheet.cell(row = row_i,column = col_j).value)
                    exe_lines.append(col_val)
                # -------- excel file read end -----------
            elif xl == 0:
                # ------------- csv file read ---------
                with open(filename) as csv_file:
                    csv_reader = csv.reader(csv_file, delimiter='\t')
                    for row in csv_reader:
                        exe_lines.append(row)
            lenchk = exe_lines[0]
            len_of_len_chk = len(lenchk)
            if len(credential) < 6:
                dataset_count = len(exe_lines[0])
            else:
                for kk in range(3,len_of_len_chk):
                    if (lenchk[kk] == None) or (lenchk[kk] == ""):
                        dataset_count = kk
                        break
                    else:
                        dataset_count = kk+1
            # print(dataset_count,"data set count")
            # exe_log.info("Removing headers - "+str(exe_lines[0])+str(dataset_count))
            exe_lines.pop(0)
                # --------------- csv file read end ------
            self.wait_till(2)
            exe_log.info(str(exe_dict)+" : Execution started ")
            exe_state = 0
            exe_dict_fun = {"id":self.byeid, "css":self.bycss, "multicss":self.bymulticss, "xpath":self.byxpath, "plink":self.byplink, "link":self.byelink, "switchframe":self.byswitchframe,
             "wait":self.py_wait,"switchalert":self.capturealert, "switchwindow": self.switchwindowto,"js": self.jsscript,"refresh": self.browser_refresh,"pid":self.bypid, "destroywindow" : self.destroywindow, 
             "open": self.openlink, "snapshot": self.saveimage, "movetoclick":self.move_to_click,"scroll":self.scrollbar,"loop":self.getloop}
            break_loop = 1
            for dataset in range(3,dataset_count):
                if break_loop == 0:
                    break
                else:
                    exe_log.info("Executing DATASET : "+str(dataset-2))
                    for ran in range(line_exe,len(exe_lines)):
                        try:
                            code_line = exe_lines[ran]
                            if code_line[0] is None:
                                exe_log.warning("Empty line or Unfilled line need to remove in Test-script")
                                continue
                            if len(code_line) == 2:
                                code_line.append(" ")
                            elif len(code_line[0]) == 0:
                                exe_log.warning("Empty line or Unfilled line need to remove in Test-script")
                                continue
                            # var = str(code_line[0])
                            code_line[0] = (str(code_line[0]).strip()).lower()
                            code_line[2] = str(code_line[2]).strip()
                            var = exe_dict_fun[code_line[0]]
                            if var != self.jsscript:
                                code_line[1] = str(code_line[1]).strip()
                            if code_line[2] == "input":
                                if (code_line[dataset] is None) or (len(str(code_line[dataset])) == 0):
                                    fr = open("Run_input.txt","w+")
                                    fr.close()
                                    for i in range(0,30):
                                        f_read = open("Run_input.txt","r")
                                        runip = (f_read.read()).split("\n")
                                        if len(runip[0]) == 0:
                                            time.sleep(2)
                                            pass
                                        elif len(runip[0]) >= 1:
                                            f_read.close()
                                            break
                                    code_line[dataset] = runip[0]
                                    exe_log.info("Run time value : '"+str(runip[0])+"' Readed")
                            var(code_line[1],code_line[2],str(code_line[dataset]))
                            if ".csv" in exe_dict:
                                if len(code_line[dataset]) == 0:
                                    code_line[dataset] = "None"
                                if len(code_line[2]) == 0:
                                    code_line[2] = "None"
                                pass
                            exe_log.info(str(code_line[0])+" : "+str(code_line[1])+" : "+str(code_line[2])+" : "+str(code_line[dataset])+" - Line executed ")
                            break_loop = 1
                        except Exception as exc:
                            exe_log.error("Error occured...Please correct the testscript")
                            exe_log.error("Exception Raised : "+str(exc.__class__)+" occured")
                            exe_state = 1
                            self.saveimage("fail_error","fail_error","fail_error")
                            self.enddriver(exe_dict,exe_state)
                            break_loop = 0
                            raise exc.__class__
                            break
                    pass
                self.wait_till(1)
                if line_exe != 0:
                    if dataset == dataset_count-1:
                        exe_log.info("Loop count Ended due to no more values")
            if break_loop != 0:
                self.enddriver(exe_dict,exe_state)
        except UnexpectedAlertPresentException as Sele_exp:
            exe_log.error(Sele_exp.__class__)
            exe_log.error("Test Execution Intialisation Error occured")
            self.saveimage("fail_error","fail_error","fail_error")
            # exe_log.error(type(err).__name__)
            self.enddriver(exe_dict,1)
            raise Sele_exp