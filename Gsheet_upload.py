import gspread
from oauth2client.service_account import ServiceAccountCredentials
from pprint import pprint
import time
import csv
import os
import time
import openpyxl
from datetime import datetime
import json

client_secret = {
  'type': 'service_account',
  'project_id': 'bot-mail-308306',
  'private_key_id': 'cd07e07295a4de8f53dcbf4b4a875f0077914cc1',
  'private_key': '-----BEGIN PRIVATE KEY-----\nMIIEvwIBADANBgkqhkiG9w0BAQEFAASCBKkwggSlAgEAAoIBAQDZ62FDcOJ6sLUD\nLPj90IbKQQXzFLlf5d1snYzYWG3azVSy1chebjYaF2xH/c1RmuABeWDYnPYhGoYT\n2gErI26/44CHm9RqOd4Nr1SZS8sv2Tanl8554lvyUezNQJBBGklFOLx1yj+x6Lb5\nccaku+ffh4HmwGgP/pFROolnVxb5G+quGBz/ooLENsqy3kMBpmEWLaoLtrvWv3RH\nyYw2forbubClq4vnOvCtpOAKuwNadVDcWCnvY/H4RLqa4xC0eiS5gmN1b8XDC5fA\nOnQi8M41vgV9tqSVOCZe+mEjnPm+78qAFcBsRaQ0+aaKLKxg0eRcV8th4FXZvlvh\nA7ir1iDxAgMBAAECggEAAdO0zkTwygiJcxNBVkIVaBnLrcBfD7bfTc7QVkFU3uP1\nlQx3yT5xTJBBfHpcI08eieER2nRjwolKDrL+a0q5Ft0zFTcUrnwyYWOoXHg1bRth\nbLNiyIt6uhWwoxsABUZfsqls/4Ek6KxgPTmUacMmbC2qGeqZChTYLX1iCNr7uqgF\naT2SjiyrkRR4y5h00nTfMBsGIo0HEkRruiTuPaNiDpubfeKI0DRPbauIb1QN3lET\nwKOBAsGPQs6OkcLnij+DvS797tTB4eixrWUVsV9FOB0cl9EkT6bu2lYyeeJ9C7K6\njz/gKUmjq9v3XoV/j5Ly6Tay6ck0LCmZ1ul6xp/ZBQKBgQDsQhJJuzTQLDkYXAFQ\nHVxmWAljl3iq1khlLcCPOOYfbBOFUW6vNHWIpxPrxzZq/++qKfv5xA+E9fNDLjE2\nt8UPSxlX8ad0gM1A0FVnROiXpMgdA4ympYz4jtqEkxEvSSqv+wZynTqB5LVRlt+H\nTNSr/gU/EdwFD1wKVJWoH88rPwKBgQDsIQQ2Q85DxKZrb33cXGGmAbSm1v6eJK50\ndj2dG3yAqnXSayjQCHGbfmD3fH7PDnJB068G/rKWJnyEycM1jEpXUbM6ZZcC1AQ9\nbNJfVIysrjYipcYe5+SdZ+OjKnkibtWZ/H8JwmaueuMAmQZ+MQBb9j2S+NWWTTUf\ndJ9kGzCXzwKBgQCT4e+Mk0/FlkjOxJhVy2sQ7L5DxoZTePmry22iuJziwkWdx5Ee\nvhlrxhkkFngrlbQwsUL13hFQyprQTFOlynXg0sM0VyDYPd88WKt8jpeAhDZ2Yshh\nb3LECrchEOVK0eOUb+5jQ6NhtwT/4w5bpG3ikog6KDWuPSUBmlxq9i45BQKBgQCC\n/X3crJjMN1cq2lqz1/vDn8abl0EAcdDlgoKDmS/kIKvj4tZtMYH0hJM+N64RJy5z\nHf9Gnhxr90W0VEkad4z2C7ileJ3hT8RaXvgFMl4nnS0i41uY4YONXoiFh5ZTZ7DP\nD6dZvqI/wKQVPM7BSmWsnlliIvA5xPLnVEbnvb2jOQKBgQDS/Gqjn8ilo4l8Y8Gr\naiyS2unlm5WCGtmlqO6IzsnXsDT+Z/WWQo4JSOOt3rLebWFLZPIPQ2Ph/byEuC3y\n4kxVTo5cSl2nUOtWC3CjijOqY8URpWhjoDT6/kNevSamjAWESYypPGitAUvOO0jz\nbA3oHV91PQbih6rH0cfsJHbhzQ==\n-----END PRIVATE KEY-----\n',
  "client_email": "bot-mail-861@bot-mail-308306.iam.gserviceaccount.com",
  "client_id": "118153670513512339504",
  "auth_uri": "https://accounts.google.com/o/oauth2/auth",
  "token_uri": "https://oauth2.googleapis.com/token",
  "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
  "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/bot-mail-861%40bot-mail-308306.iam.gserviceaccount.com"
}

with open("client_secret.json", "w") as outfile:
    json.dump(client_secret, outfile)
# use creds to create a client to interact with the Google Drive API
scope = [
'https://www.googleapis.com/auth/spreadsheets',
'https://www.googleapis.com/auth/drive'
]


creds = ServiceAccountCredentials.from_json_keyfile_name('client_secret.json', scope)
# creds = ServiceAccountCredentials.from(client_secret,scope)
client = gspread.authorize(creds)

ff = open("client_secret.json","w+")
ff.close()

# Find a spreadsheet by name
# Make sure you use the right name here.
# sheet = client.open("test py sele")

# Access the worksheet by name 
# get_sheet = sheet.worksheet("Sheet2")
# sheet = client.open("My Anon mail")

# Extract and print all of the values
# list_of_hashes = sheet.get_all_records()
# print(list_of_hashes)

#  To add a worksheet in spreadsheet 
# worksheet = sheet.add_worksheet(title="A worksheet", rows="100", cols="20") 

# if you want you can define row and columns of the sheet
#  worksheet = sheet.add_worksheet(title="A worksheet") 

# To get the list of worksheet present in the spreadsheet
# worksheet_list = sheet.worksheets()

# To print all the values in the sheet on pretty print in list of list format
# pprint(worksheet.get_all_values())
# To print all the values in the sheet on pretty print in list of dictionary format
# pprint(worksheet.get_all_records())

# To add rows in the spreadsheet
# sheet.add_rows(20000)

# To insert row in worksheet define worksheet variable and do insert
# worksheet.insert_row(row,i)

# To update a range of cells define worksheet in worksheet variable and do update
# worksheet.update('A1:B2', [[1, 2], [3, 4]])
# worksheet.update_cell(1, 2, 'Bingo!')  - to update a cell 1 is row and 2 is column or 
# worksheet.update('B1', 'Bingo!') - to update a cell 

row = ["haii this is gsheet","abcd","haii"]

def sheet_create(sheet_name):
    main_sheet.add_worksheet(title=sheet_name,rows=1000,cols=100)
    time.sleep(2)
# def row_insert(values,sheet_name):
#     for i in range(1,len(values)):
#         time.sleep(2)
#         sheet_name.insert_row(values[i-1],i)
def cell_update(values,sheet_name):
    sheet_name.update(values)
    # for i in range(1,i_row):
    #     # sheet.update_cell(i, 1, "I just wrote to a spreadsheet using Python! 1")
    #     time.sleep(4)
    #     i_val = values[i-1]
    #     for j in range(1,j_col):
    #     # get_sheet.insert_row(row,i)
    #         sheet_name.update_cell(i,j,i_val[j-1])
main_sheet = ""
def sheet_detail():
    global main_sheet
    print("Make sure the changes to gsheet.txt and give the Google sheet name you want to ACCESS...")
    gsheet = open("gsheet.txt","r")
    gsheet_name = (gsheet.read()).split(" - ")
    gsheet.close()
    print("Accessing gsheet.txt file...")
    print("Opening the Gsheet :"+str(gsheet_name[1]))
    main_sheet = client.open(gsheet_name[1])
    
    # Access the file list name from run_cases.csv 
    file_csv = open("Run_cases.csv","r",encoding="utf")
    list_files = (file_csv.read()).split("\n")
    file_csv.close()
    fileDir = os.path.dirname(os.path.realpath("__file__"))
    total_row = ""
    total_col = ""
    list_files.pop(0)
    list_files.pop(len(list_files)-1)
    for sheet_name in list_files:  
        worksheet_list = main_sheet.worksheets()
        wk_list = []
        for wk in worksheet_list:
            wk_list.append((str(wk).split(":"))[0])
        exe_lines = []
        if ".csv" in sheet_name:
            filename = os.path.join(fileDir, "casefolder\\"+sheet_name)
            xl = 0
        elif ".xlsx" in sheet_name:
            filename = os.path.join(fileDir, "casefolder\\"+str(sheet_name))
            xl = 1
        else:
            filename = os.path.join(fileDir, "casefolder\\"+str(sheet_name)+".xlsx")
            xl = 1
        if xl == 1:
            # -------- excel file read ----------- data store in exe_lines variable
            excel_book = openpyxl.load_workbook(filename)
            excel_sheet = excel_book.active
            # To define the total row and total col count
            total_row = excel_sheet.max_row+1
            total_col = excel_sheet.max_column+1
            for row_i in range(1,total_row):
                col_val = []
                for col_j in range(1,total_col):
                    col_val.append(excel_sheet.cell(row = row_i,column = col_j).value)
                exe_lines.append(col_val)
            # -------- excel file read end -----------
        elif xl == 0:
            # ------------- csv file read --------- data store in exe_lines variable
            with open(filename) as csv_file:
                csv_reader = csv.reader(csv_file, delimiter='\t')
                for row in csv_reader:
                    exe_lines.append(row)
            # --------------- csv file read end ------ 
        sheet_name = sheet_name.split(".")
        sheet_name_conf =  "<Worksheet '"+sheet_name[0]+"' id"
        sheet_name = sheet_name[0]
        new_sheet = 0
        for wk_name in wk_list:
            # print(sheet_name_conf,wk_name)
            if sheet_name_conf == wk_name:
                new_sheet = 1
        if new_sheet == 1:
            print("Exisiting worksheet : "+sheet_name+" is found")
            exe_sheet = main_sheet.worksheet(sheet_name)
            print("Sheet Name : "+sheet_name+" is UPDATED by given values...")
            cell_update(exe_lines,exe_sheet)
            # row_insert(exe_lines,exe_sheet)
        elif new_sheet == 0:
            exe_sheet = sheet_name
            sheet_create(exe_sheet)
            print("New worksheet : "+sheet_name+" created in Google sheet")
            exe_sheet = main_sheet.worksheet(sheet_name)
            # row_insert(exe_lines,exe_sheet)
            cell_update(exe_lines,exe_sheet)
            print("Sheet Name : "+sheet_name+" is INSERTED by given values...")
sheet_detail()